from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font

try:
    from reportlab.lib.pagesizes import letter
    from reportlab.pdfgen import canvas
except Exception:
    canvas = None


def export_billing_xlsx(path: str, project: dict, line_items: list[dict]):
    """
    line_items: list of dicts with keys:
      sample_name, test_code, test_name, cost
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Billing"

    ws["A1"] = "Project"
    ws["B1"] = project.get("job_name", "")
    ws["A2"] = "File #"
    ws["B2"] = project.get("file_number", "")
    ws["A3"] = "Client"
    ws["B3"] = project.get("client_type", "")
    ws["A4"] = "Billing Rate"
    ws["B4"] = project.get("billing_rate_id", "")

    header_row = 6
    headers = ["Sample (Depth)", "Test Code", "Test Name", "Cost"]
    for col, val in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col, value=val)
        cell.font = Font(bold=True)

    total = 0.0
    row = header_row + 1
    for item in line_items:
        sample = item.get("sample_name", "")
        depth = item.get("depth_raw", "")
        label = f"{sample} @ {depth}" if depth else sample
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=item.get("test_code", ""))
        ws.cell(row=row, column=3, value=item.get("test_name", ""))
        cost = float(item.get("cost", 0.0) or 0.0)
        ws.cell(row=row, column=4, value=cost)
        total += cost
        row += 1

    ws.cell(row=row + 1, column=3, value="Total").font = Font(bold=True)
    ws.cell(row=row + 1, column=4, value=total).font = Font(bold=True)

    Path(path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def export_billing_pdf(path: str, project: dict, line_items: list[dict]):
    if canvas is None:
        raise RuntimeError("PDF export requires reportlab. Please install dependencies.")

    c = canvas.Canvas(path, pagesize=letter)
    width, height = letter

    y = height - 50
    c.setFont("Helvetica-Bold", 14)
    c.drawString(50, y, "TCI Laboratory Billing Summary")

    y -= 25
    c.setFont("Helvetica", 10)
    c.drawString(50, y, f"Project: {project.get('job_name', '')}")
    y -= 15
    c.drawString(50, y, f"File #: {project.get('file_number', '')}")
    y -= 15
    c.drawString(50, y, f"Client: {project.get('client_type', '')}")
    y -= 15
    c.drawString(50, y, f"Billing Rate: {project.get('billing_rate_id', '')}")

    y -= 25
    c.setFont("Helvetica-Bold", 10)
    c.drawString(50, y, "Sample (Depth)")
    c.drawString(200, y, "Test")
    c.drawString(420, y, "Cost")
    y -= 12
    c.setFont("Helvetica", 10)

    total = 0.0
    for item in line_items:
        if y < 60:
            c.showPage()
            y = height - 50
        sample = item.get("sample_name", "")
        depth = item.get("depth_raw", "")
        label = f"{sample} @ {depth}" if depth else sample
        c.drawString(50, y, label)
        c.drawString(200, y, item.get("test_name", ""))
        cost = float(item.get("cost", 0.0) or 0.0)
        c.drawRightString(470, y, f"{cost:.2f}")
        total += cost
        y -= 12

    y -= 10
    c.setFont("Helvetica-Bold", 10)
    c.drawString(200, y, "Total")
    c.drawRightString(470, y, f"{total:.2f}")

    Path(path).parent.mkdir(parents=True, exist_ok=True)
    c.save()
