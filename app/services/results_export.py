from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, landscape
    from reportlab.pdfgen import canvas as pdf_canvas
    from reportlab.platypus import Table, TableStyle
except Exception:
    pdf_canvas = None

PDF_HEADER_BLUE = "D8EAF9"
PDF_SUBHEADER_BLUE = "EDF5FD"
PDF_GRID_BLUE = "9EBBD8"


def export_results_matrix_xlsx(path: str, project: dict, samples: list[dict], tests: list[dict], results: list[dict]):
    schema = _build_schema(tests, results)
    row_map = _row_lookup(results)

    wb = Workbook()
    ws = wb.active
    ws.title = "Test Results"

    total_cols = 3 + len(schema)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    ws.cell(row=1, column=1, value=project.get("job_name", ""))
    ws.cell(row=1, column=1).font = Font(size=14, bold=True)
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="center")

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=total_cols)
    ws.cell(row=2, column=1, value=f"File Number: {project.get('file_number', '')}")
    ws.cell(row=2, column=1).font = Font(size=11, bold=True)
    ws.cell(row=2, column=1).alignment = Alignment(horizontal="center")

    # Row 4: ASTM/designation, Row 5: subcolumn name + unit
    ws.cell(row=4, column=1, value="Sample Location")
    ws.cell(row=4, column=2, value="Sample Depth")
    ws.cell(row=4, column=3, value="Sample Type")
    for col in (1, 2, 3):
        ws.cell(row=4, column=col).font = Font(size=9, bold=True)
        ws.cell(row=4, column=col).fill = PatternFill("solid", fgColor=PDF_HEADER_BLUE)
        ws.cell(row=4, column=col).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.cell(row=5, column=col).font = Font(size=8, bold=True)
        ws.cell(row=5, column=col).fill = PatternFill("solid", fgColor=PDF_SUBHEADER_BLUE)
        ws.cell(row=5, column=col).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 12

    for i, col_meta in enumerate(schema, start=4):
        ws.cell(row=4, column=i, value=col_meta["designation"])
        ws.cell(row=4, column=i).font = Font(size=8, bold=True)
        ws.cell(row=4, column=i).fill = PatternFill("solid", fgColor=PDF_HEADER_BLUE)
        ws.cell(row=4, column=i).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        sub = col_meta["label"]
        if col_meta["unit"]:
            sub = f"{sub}\n[{col_meta['unit']}]"
        ws.cell(row=5, column=i, value=sub)
        ws.cell(row=5, column=i).font = Font(size=8, bold=True)
        ws.cell(row=5, column=i).fill = PatternFill("solid", fgColor=PDF_SUBHEADER_BLUE)
        ws.cell(row=5, column=i).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[_col_letter(i)].width = 18

    # Merge ASTM/designation headers so each test block has one top heading.
    _merge_grouped_headers_xlsx(ws, schema, start_col=4, header_row=4)

    row_idx = 6
    for sample in samples:
        loc, s_type = _sample_location_and_type(sample)
        ws.cell(row=row_idx, column=1, value=loc)
        ws.cell(row=row_idx, column=2, value=sample.get("depth_raw") or "")
        ws.cell(row=row_idx, column=3, value=s_type)
        ws.cell(row=row_idx, column=1).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.cell(row=row_idx, column=2).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.cell(row=row_idx, column=3).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for col_idx, col_meta in enumerate(schema, start=4):
            raw = row_map.get((sample["id"], col_meta["test_id"]), {})
            text = col_meta["extractor"](raw)
            ws.cell(row=row_idx, column=col_idx, value=text)
            ws.cell(row=row_idx, column=col_idx).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        row_idx += 1

    ws.freeze_panes = "D6"
    Path(path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def export_results_matrix_pdf(path: str, project: dict, samples: list[dict], tests: list[dict], results: list[dict]):
    if pdf_canvas is None:
        raise RuntimeError("PDF export requires reportlab. Please install dependencies.")

    schema = _build_schema(tests, results)
    row_map = _row_lookup(results)
    if not schema:
        raise RuntimeError("No entered test data to export.")

    header1 = ["Sample Location", "Sample Depth", "Sample Type"] + [c["designation"] for c in schema]
    header2 = ["", "", ""]
    for c in schema:
        txt = c["label"]
        if c["unit"]:
            txt = f"{txt} [{c['unit']}]"
        header2.append(txt)

    data = [header1, header2]
    for sample in samples:
        loc, s_type = _sample_location_and_type(sample)
        row_vals = [loc, sample.get("depth_raw") or "", s_type]
        for c in schema:
            raw = row_map.get((sample["id"], c["test_id"]), {})
            row_vals.append(c["extractor"](raw))
        data.append(row_vals)

    page_w, page_h = landscape(letter)
    margin = 18
    title_gap = 34
    available_w = page_w - 2 * margin
    available_h = page_h - (2 * margin) - title_gap

    fixed_left = [84, 72, 60]
    test_col_base = 104
    col_widths = fixed_left + [test_col_base] * len(schema)

    table = Table(data, colWidths=col_widths, repeatRows=2)
    style_cmds = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(f"#{PDF_HEADER_BLUE}")),
        ("BACKGROUND", (0, 1), (-1, 1), colors.HexColor(f"#{PDF_SUBHEADER_BLUE}")),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor(f"#{PDF_GRID_BLUE}")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("FONTNAME", (0, 0), (-1, 1), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 1), 7),
        ("FONTSIZE", (0, 2), (-1, -1), 7.25),
        ("TOPPADDING", (0, 0), (-1, -1), 2),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
    ]
    _apply_grouped_header_spans_pdf(style_cmds, data, schema, start_col=3)
    table.setStyle(TableStyle(style_cmds))

    Path(path).parent.mkdir(parents=True, exist_ok=True)
    c = pdf_canvas.Canvas(path, pagesize=landscape(letter))

    tw, th = table.wrapOn(c, available_w, available_h)
    scale = min(available_w / max(tw, 1), available_h / max(th, 1))
    if scale <= 0:
        scale = 1.0

    draw_w = tw * scale
    draw_h = th * scale
    x = margin + (available_w - draw_w) / 2

    # Top-anchor the table instead of vertical-centering.
    title_gap = 8
    subtitle_gap = 12
    top_block = title_gap + subtitle_gap + 16
    top_y = page_h - margin - top_block
    y = max(margin, top_y - draw_h)

    # Keep title block visually attached to the table, not fixed at page top.
    title_y = y + draw_h + title_gap + subtitle_gap
    subtitle_y = y + draw_h + title_gap

    title = project.get("job_name", "")
    subtitle = f"File Number: {project.get('file_number', '')}"
    logo_path = Path(__file__).resolve().parents[2] / "assets" / "terrapacific_logo.png"
    if logo_path.exists():
        try:
            # Draw logo above title block if available.
            logo_w = 160
            logo_h = 40
            c.drawImage(str(logo_path), (page_w - logo_w) / 2, title_y + 6, width=logo_w, height=logo_h, preserveAspectRatio=True, mask="auto")
            title_y -= 4
            subtitle_y -= 4
        except Exception:
            pass
    c.setFont("Helvetica-Bold", 13)
    c.setFillColor(colors.HexColor("#2F86DE"))
    c.drawCentredString(page_w / 2, title_y, title)
    c.setFont("Helvetica", 9)
    c.setFillColor(colors.HexColor("#15385B"))
    c.drawCentredString(page_w / 2, subtitle_y, subtitle)

    c.saveState()
    c.translate(x, y)
    c.scale(scale, scale)
    table.drawOn(c, 0, 0)
    c.restoreState()
    c.showPage()
    c.save()


def _build_schema(tests: list[dict], results: list[dict]):
    # map test_id -> list of rows for optional column visibility
    by_test = {}
    for t in tests:
        by_test[t["id"]] = [r for r in results if r["test_id"] == t["id"]]

    schema = []
    for test in tests:
        name = test["name"]
        rows = by_test.get(test["id"], [])
        if not _has_entered_result(rows):
            continue

        if name == "Chem":
            has_r = _any_value(rows, "result_value3") or _any_value(rows, "result_value4")
            has_ph = _any_value(rows, "result_value4")
            if has_r:
                schema.append(_col(test["id"], "Corrosivity Series", "CTM643 Resistivity", "ohm-cm", _chem_resistivity))
            schema.append(_col(test["id"], "Corrosivity Series", "CTM422 Chloride Content", "%", _chem_chloride))
            schema.append(_col(test["id"], "Corrosivity Series", "CTM417 Sulfate Content", "%", _chem_sulfate))
            if has_ph:
                schema.append(_col(test["id"], "Corrosivity Series", "pH", "pH", _chem_ph))
            continue

        if name == "Sieve Part. Analysis":
            schema.append(_col(test["id"], "ASTM D 422", "USCS Classif.", "", _sieve_uscs))
            schema.append(_col(test["id"], "ASTM D 422", "Passing No. 200", "%", _value2))
            continue

        if name == "Sand Cone":
            schema.append(_col(test["id"], "ASTM D 1556", "Dry Density", "pcf", _value1))
            schema.append(_col(test["id"], "ASTM D 1556", "Moisture Content", "%", _value2))
            continue

        if name in ("Max Density", "C Max"):
            schema.append(_col(test["id"], "ASTM D 1557", "Maximum Dry Density", "pcf", _value1))
            schema.append(_col(test["id"], "ASTM D 1557", "Opt. Moisture Content", "%", _value2))
            continue

        if name in ("698 Max",):
            schema.append(_col(test["id"], "ASTM D 698", "Maximum Dry Density", "pcf", _value1))
            schema.append(_col(test["id"], "ASTM D 698", "Opt. Moisture Content", "%", _value2))
            continue

        if name in ("Moisture Content",):
            schema.append(_col(test["id"], "ASTM D 2216", "Moisture Content", "%", _value1))
            continue

        if name in ("R-Value", "R-Value by Equilibration"):
            schema.append(_col(test["id"], "ASTM D 2844", "R-Value by Equilibration", "", _value1))
            continue

        if name in ("Field Density/Moisture",):
            schema.append(_col(test["id"], "ASTM D 2937", "Dry Density", "pcf", _value1))
            schema.append(_col(test["id"], "ASTM D 2937", "Moisture Content", "%", _value2))
            if _any_value(rows, "result_value3"):
                schema.append(_col(test["id"], "ASTM D 2937", "Saturation", "%", _value3))
            elif _any_nonempty_str(rows, "result_unit3"):
                schema.append(_col(test["id"], "ASTM D 2937", "Saturation", "%", _value3))
            continue

        if name == "Expansion Index":
            schema.append(_col(test["id"], "ASTM D 4829", "Index", "EI", _value1))
            schema.append(_col(test["id"], "ASTM D 4829", "Potential", "", _ei_potential))
            continue

        if name == "Direct Shear":
            schema.append(_col(test["id"], "ASTM D 3080", "Peak (PHI)", "deg", _value1))
            schema.append(_col(test["id"], "ASTM D 3080", "Peak (c)", "psf", _value2))
            continue

        if name in ("LL/PL", "Atterberg Limits"):
            schema.append(_col(test["id"], "ASTM D 4318", "Liquid Limit", "%", _value1))
            schema.append(_col(test["id"], "ASTM D 4318", "Plasticity Index", "%", _value2))
            continue

        if name in ("Swell/Hydro", "Hydro Response"):
            schema.append(_col(test["id"], "ASTM D 4546", "Hydro Response", "%", _value1))
            schema.append(_col(test["id"], "ASTM D 4546", "Normal Stress", "psf", _value2))
            continue

        # fallback: single column
        schema.append(_col(test["id"], test.get("code", name), name, _unit_from_results(rows), _value1))
    return schema


def _col(test_id, designation, label, unit, extractor):
    return {
        "test_id": test_id,
        "designation": designation,
        "label": label,
        "unit": unit,
        "extractor": extractor,
    }


def _row_lookup(results: list[dict]):
    lookup = {}
    for r in results:
        lookup[(r["sample_id"], r["test_id"])] = r
    return lookup


def _value1(row):
    return _fmt_plain(row.get("result_value"))


def _value2(row):
    return _fmt_plain(row.get("result_value2"))


def _value3(row):
    return _fmt_plain(row.get("result_value3"))


def _value4(row):
    return _fmt_plain(row.get("result_value4"))


def _chem_resistivity(row):
    # all-four mode stores resistivity in value1
    has_all_four = row.get("result_value3") is not None or row.get("result_value4") is not None
    if not has_all_four:
        return ""
    return _fmt_plain(row.get("result_value"))


def _chem_chloride(row):
    # default mode uses value1 for chloride; all-four uses value3
    has_all_four = row.get("result_value3") is not None or row.get("result_value4") is not None
    if has_all_four:
        return _fmt_plain(row.get("result_value3"))
    return _fmt_plain(row.get("result_value"))


def _chem_sulfate(row):
    return _fmt_plain(row.get("result_value2"))


def _chem_ph(row):
    return _fmt_plain(row.get("result_value4"))


def _ei_potential(row):
    # Potential is stored as text in result_unit2 by current results entry flow.
    potential = row.get("result_unit2")
    if isinstance(potential, str) and potential.strip():
        return potential.strip()
    return _fmt_plain(row.get("result_value2"))


def _sieve_uscs(row):
    # If USCS is captured in unit text, show it.
    unit = row.get("result_unit")
    if unit:
        return str(unit)
    return _fmt_plain(row.get("result_value"))


def _fmt_plain(value):
    if value is None:
        return ""
    try:
        txt = f"{float(value):.1f}"
    except Exception:
        txt = str(value)
    return txt


def _sample_location_and_type(sample: dict):
    name = (sample.get("sample_name") or "").strip()
    sample_type = (sample.get("sample_type") or "").strip().upper()
    upper = name.upper()
    # Backward-compatible fallback for older records without explicit sample_type.
    if sample_type in {"SB", "MB", "LB", "SPT", "(RING)", "RING", "RING)"}:
        if sample_type in {"(RING)", "RING", "RING)"}:
            sample_type = "Ring"
        return name, sample_type
    if upper.startswith("B-"):
        sample_type = "B"
    elif upper.startswith("T-"):
        sample_type = "T"
    elif upper.startswith("HA-"):
        sample_type = "HA"
    elif upper.startswith("C-"):
        sample_type = "C"
    return name, sample_type


def _unit_from_results(rows):
    for r in rows:
        for key in ("result_unit", "result_unit2", "result_unit3", "result_unit4"):
            if r.get(key):
                return str(r[key])
    return ""


def _any_value(rows, key):
    for r in rows:
        if r.get(key) is not None:
            return True
    return False


def _any_nonempty_str(rows, key):
    for r in rows:
        val = r.get(key)
        if isinstance(val, str) and val.strip():
            return True
    return False


def _has_entered_result(rows):
    for r in rows:
        for key in ("result_value", "result_value2", "result_value3", "result_value4"):
            if r.get(key) is not None:
                return True
        for key in ("result_unit", "result_unit2", "result_unit3", "result_unit4"):
            val = r.get(key)
            if isinstance(val, str) and val.strip():
                return True
    return False


def _col_letter(col_index: int) -> str:
    letters = ""
    idx = col_index
    while idx > 0:
        idx, rem = divmod(idx - 1, 26)
        letters = chr(65 + rem) + letters
    return letters


def _merge_grouped_headers_xlsx(ws, schema, start_col, header_row):
    i = 0
    while i < len(schema):
        start = i
        des = schema[i]["designation"]
        while i + 1 < len(schema) and schema[i + 1]["designation"] == des:
            i += 1
        end = i
        if end > start:
            c1 = start_col + start
            c2 = start_col + end
            ws.merge_cells(start_row=header_row, start_column=c1, end_row=header_row, end_column=c2)
            ws.cell(row=header_row, column=c1, value=des)
            ws.cell(row=header_row, column=c1).font = Font(size=8, bold=True)
            ws.cell(row=header_row, column=c1).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        i += 1


def _apply_grouped_header_spans_pdf(style_cmds, data, schema, start_col):
    # row 0 in PDF data is designation header
    i = 0
    while i < len(schema):
        start = i
        des = schema[i]["designation"]
        while i + 1 < len(schema) and schema[i + 1]["designation"] == des:
            i += 1
        end = i
        if end > start:
            c1 = start_col + start
            c2 = start_col + end
            style_cmds.append(("SPAN", (c1, 0), (c2, 0)))
            for j in range(start + 1, end + 1):
                data[0][start_col + j] = ""
            data[0][c1] = des
        i += 1
